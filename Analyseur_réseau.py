# -*- coding: utf-8 -*-
"""
SAE 1.05 - Analyseur TKINTER avec GRAPHIQUES + EXPORT EXCEL
Version finale : fenêtre avec graphiques + export Excel avec camemberts
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from collections import Counter
import csv
import re
from datetime import datetime
import zipfile
import io

# Matplotlib pour les graphiques dans la fenêtre
try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

# ============================================================================
# FONCTION D'ANALYSE
# ============================================================================

def analyser_fichier(chemin):
    """Analyse le fichier et retourne les stats."""
    
    with open(chemin, "r", encoding="utf-8", errors="ignore") as f:
        lignes = [l.strip() for l in f if l.strip() and not l.strip().startswith("0x")]
    
    sources = []
    destinations = []
    flags = []
    
    for ligne in lignes:
        parts = ligne.split()
        if ">" in parts:
            idx = parts.index(">")
            if 0 < idx < len(parts) - 1:
                sources.append(parts[idx - 1].strip())
                destinations.append(parts[idx + 1].rstrip(" :,").strip())
        if "Flags" in parts:
            i = parts.index("Flags")
            if i < len(parts) - 1:
                flags.append(parts[i + 1].strip())
    
    cnt_src = Counter(sources)
    cnt_dst = Counter(destinations)
    cnt_flags = Counter(flags)
    
    err_regex = re.compile(r"\b(ERROR|Error|ERR|Exception|CRITICAL|FATAL|failed|failure|denied)\b", re.IGNORECASE)
    erreurs = []
    err_types = Counter()
    
    for idx, ligne in enumerate(lignes, 1):
        m = err_regex.search(ligne)
        if m:
            t = m.group(0).upper()
            err_types[t] += 1
            erreurs.append({"ligne": idx, "type": t, "msg": ligne[:150]})
    
    alertes = []
    
    for ip, count in cnt_dst.most_common(5):
        if count > 50:
            alertes.append(f"🔴 DOS : {count} connexions vers {ip}")
    
    syn_total = sum(c for flag, c in cnt_flags.items() if "[S]" in flag)
    if syn_total > 50:
        alertes.append(f"🟠 SYN FLOOD : {syn_total} paquets SYN")
    
    dest_map = dict(cnt_dst)
    for ip, env in cnt_src.most_common(10):
        rec = dest_map.get(ip, 0)
        ratio = env / max(rec, 1)
        if env > 20 and ratio > 5:
            alertes.append(f"🟡 DESEQUILIBRE : {ip} ({env}→{rec}, ratio {ratio:.1f}:1)")
    
    if erreurs:
        alertes.append(f"❌ ERREURS : {len(erreurs)} lignes d'erreur")
    
    total_src = sum(cnt_src.values())
    total_dst = sum(cnt_dst.values())
    total_flags = sum(cnt_flags.values())
    
    return {
        "lignes": len(lignes),
        "sources": cnt_src.most_common(10),
        "destinations": cnt_dst.most_common(10),
        "flags": list(cnt_flags.items()),
        "erreurs": erreurs,
        "err_types": dict(err_types),
        "alertes": alertes,
        "total_src": total_src,
        "total_dst": total_dst,
        "total_flags": total_flags,
    }

# ============================================================================
# INTERFACE GRAPHIQUE AVEC GRAPHIQUES
# ============================================================================

class AnalyseurApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🛡️ SAE 1.05 - Analyseur Réseau avec Graphiques")
        self.root.geometry("1400x800")
        self.root.configure(bg="#2c3e50")
        
        self.stats = None
        self.fichier = None
        
        # Frame principal avec scrollbar
        main_canvas = tk.Canvas(root, bg="#2c3e50")
        scrollbar = tk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
        self.scrollable_frame = tk.Frame(main_canvas, bg="#2c3e50")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Titre
        titre = tk.Label(
            self.scrollable_frame, 
            text="🛡️ Analyseur SAE 1.05 - Avec Graphiques", 
            font=("Arial", 24, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        titre.pack(pady=20)
        
        # Bouton upload
        self.btn_upload = tk.Button(
            self.scrollable_frame,
            text="📂 Sélectionner un fichier",
            font=("Arial", 14, "bold"),
            bg="#3498db",
            fg="white",
            padx=30,
            pady=15,
            command=self.upload_fichier,
            cursor="hand2"
        )
        self.btn_upload.pack(pady=10)
        
        # Frame pour graphiques
        self.graph_frame = tk.Frame(self.scrollable_frame, bg="#2c3e50")
        
        # Zone de texte résumé
        self.result_frame = tk.Frame(self.scrollable_frame, bg="#ecf0f1", relief=tk.RAISED, bd=2)
        
        scroll_frame = tk.Frame(self.result_frame, bg="#ecf0f1")
        scroll_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        txt_scrollbar = tk.Scrollbar(scroll_frame)
        txt_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.txt_result = tk.Text(
            scroll_frame,
            wrap=tk.WORD,
            yscrollcommand=txt_scrollbar.set,
            font=("Consolas", 9),
            bg="white",
            fg="#2c3e50",
            height=15
        )
        self.txt_result.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        txt_scrollbar.config(command=self.txt_result.yview)
        
        # Boutons export
        self.export_frame = tk.Frame(self.scrollable_frame, bg="#2c3e50")
        
        self.btn_excel = tk.Button(
            self.export_frame,
            text="📊 Export Excel (avec graphiques)",
            font=("Arial", 12, "bold"),
            bg="#2ecc71",
            fg="white",
            padx=20,
            pady=10,
            command=self.export_excel
        )
        self.btn_excel.pack(side=tk.LEFT, padx=5)
        
        self.btn_md = tk.Button(
            self.export_frame,
            text="📝 Export Markdown",
            font=("Arial", 12, "bold"),
            bg="#f39c12",
            fg="white",
            padx=20,
            pady=10,
            command=self.export_md
        )
        self.btn_md.pack(side=tk.LEFT, padx=5)
        
        # Message initial
        msg = tk.Label(
            self.scrollable_frame,
            text="👉 Clique sur 'Sélectionner un fichier' pour commencer\nFormats : .txt, .log, .csv, .dump",
            font=("Arial", 12),
            bg="#ecf0f1",
            fg="#2c3e50",
            pady=20
        )
        msg.pack(pady=20, fill=tk.X)
    
    def upload_fichier(self):
        chemin = filedialog.askopenfilename(
            title="Sélectionner un fichier",
            filetypes=[
                ("Fichiers texte", "*.txt"),
                ("Fichiers log", "*.log"),
                ("Fichiers CSV", "*.csv"),
                ("Fichiers dump", "*.dump"),
                ("Tous les fichiers", "*.*")
            ]
        )
        
        if not chemin:
            return
        
        self.fichier = Path(chemin)
        
        try:
            self.stats = analyser_fichier(self.fichier)
            self.afficher_resultats()
            if MATPLOTLIB_OK:
                self.afficher_graphiques()
            self.graph_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            self.result_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            self.export_frame.pack(pady=10)
            messagebox.showinfo("✅ Succès", "Analyse terminée !")
        except Exception as e:
            messagebox.showerror("❌ Erreur", f"Impossible d'analyser :\n{e}")
    
    def afficher_graphiques(self):
        """Affiche les graphiques (camemberts + barres)."""
        for widget in self.graph_frame.winfo_children():
            widget.destroy()
        
        s = self.stats
        
        # Figure avec 4 graphiques
        fig = Figure(figsize=(14, 10), facecolor='#ecf0f1')
        
        # 1. Sources (Top 5) - Camembert
        ax1 = fig.add_subplot(2, 2, 1)
        if s['sources']:
            labels = [ip[:20] for ip, _ in s['sources'][:5]]
            values = [c for _, c in s['sources'][:5]]
            colors = ['#3498db', '#2ecc71', '#f39c12', '#e74c3c', '#9b59b6']
            explode = (0.1, 0, 0, 0, 0)
            
            ax1.pie(values, labels=labels, autopct='%1.1f%%', colors=colors, 
                   explode=explode, shadow=True, startangle=90)
            ax1.set_title('Top 5 Sources (%)', fontsize=14, fontweight='bold')
        
        # 2. Destinations (Top 5) - Camembert
        ax2 = fig.add_subplot(2, 2, 2)
        if s['destinations']:
            labels = [ip[:20] for ip, _ in s['destinations'][:5]]
            values = [c for _, c in s['destinations'][:5]]
            colors = ['#e74c3c', '#f39c12', '#3498db', '#2ecc71', '#9b59b6']
            explode = (0.1, 0, 0, 0, 0)
            
            ax2.pie(values, labels=labels, autopct='%1.1f%%', colors=colors,
                   explode=explode, shadow=True, startangle=90)
            ax2.set_title('Top 5 Destinations (%)', fontsize=14, fontweight='bold')
        
        # 3. Flags TCP - Barres horizontales
        ax3 = fig.add_subplot(2, 2, 3)
        if s['flags']:
            flags_labels = [f for f, _ in s['flags'][:6]]
            flags_values = [c for _, c in s['flags'][:6]]
            
            ax3.barh(flags_labels, flags_values, color='#3498db')
            ax3.set_xlabel('Nombre de paquets', fontweight='bold')
            ax3.set_title('Flags TCP (Top 6)', fontsize=14, fontweight='bold')
            ax3.grid(axis='x', alpha=0.3)
        
        # 4. Types d'erreurs - Camembert
        ax4 = fig.add_subplot(2, 2, 4)
        if s['err_types']:
            labels = list(s['err_types'].keys())
            values = list(s['err_types'].values())
            colors = ['#e74c3c', '#f39c12', '#e67e22', '#c0392b']
            
            ax4.pie(values, labels=labels, autopct='%1.1f%%', colors=colors,
                   shadow=True, startangle=90)
            ax4.set_title('Types d\'erreurs (%)', fontsize=14, fontweight='bold')
        else:
            ax4.text(0.5, 0.5, '✅ Aucune erreur', 
                    ha='center', va='center', fontsize=16, color='#2ecc71')
            ax4.set_title('Types d\'erreurs', fontsize=14, fontweight='bold')
            ax4.axis('off')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def afficher_resultats(self):
        """Affiche le résumé texte."""
        s = self.stats
        self.txt_result.delete("1.0", tk.END)
        
        self.txt_result.insert(tk.END, "═" * 120 + "\n")
        self.txt_result.insert(tk.END, f"  📊 RÉSUMÉ - {self.fichier.name}\n")
        self.txt_result.insert(tk.END, "═" * 120 + "\n\n")
        
        self.txt_result.insert(tk.END, f"📈 Lignes : {s['lignes']} | Sources : {len(s['sources'])} | Destinations : {len(s['destinations'])} | ")
        self.txt_result.insert(tk.END, f"Erreurs : {len(s['erreurs'])} | Alertes : {len(s['alertes'])}\n\n")
        
        # Sources avec %
        self.txt_result.insert(tk.END, "🔵 TOP SOURCES\n")
        for i, (ip, c) in enumerate(s['sources'], 1):
            pct = round(c / s['total_src'] * 100, 1) if s['total_src'] > 0 else 0
            self.txt_result.insert(tk.END, f"  {i:2d}. {ip:40s} : {c:6d} ({pct:5.1f}%)\n")
        
        self.txt_result.insert(tk.END, "\n🔴 TOP DESTINATIONS\n")
        for i, (ip, c) in enumerate(s['destinations'], 1):
            pct = round(c / s['total_dst'] * 100, 1) if s['total_dst'] > 0 else 0
            self.txt_result.insert(tk.END, f"  {i:2d}. {ip:40s} : {c:6d} ({pct:5.1f}%)\n")
        
        if s['alertes']:
            self.txt_result.insert(tk.END, "\n⚠️ ALERTES\n")
            for a in s['alertes']:
                self.txt_result.insert(tk.END, f"  • {a}\n")
        
        self.txt_result.insert(tk.END, "\n" + "═" * 120 + "\n")
    
    def export_excel(self):
        """Export Excel avec graphiques intégrés (camemberts + barres)."""
        if not self.stats:
            return
        
        # Vérifie que openpyxl est installé
        try:
            from openpyxl import Workbook
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            reponse = messagebox.askyesno(
                "⚠️ Module manquant",
                "openpyxl n'est pas installé.\n\nInstaller maintenant ?"
            )
            if reponse:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("✅", "openpyxl installé ! Relance l'export.")
            return
        
        chemin = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
            initialfile="analyse_sae105.xlsx"
        )
        
        if not chemin:
            return
        
        s = self.stats
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Styles
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center = Alignment(horizontal="center", vertical="center")
        
        # ===== ONGLET 1 : RÉSUMÉ =====
        ws_resume = wb.create_sheet("📊 Résumé", 0)
        ws_resume.column_dimensions['A'].width = 30
        ws_resume.column_dimensions['B'].width = 40
        
        ws_resume.append(["Indicateur", "Valeur"])
        for cell in ws_resume[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        ws_resume.append(["📁 Fichier", self.fichier.name])
        ws_resume.append(["📅 Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_resume.append(["📄 Lignes analysées", s['lignes']])
        ws_resume.append(["🔵 Sources distinctes", len(s['sources'])])
        ws_resume.append(["🔴 Destinations distinctes", len(s['destinations'])])
        ws_resume.append(["❌ Erreurs détectées", len(s['erreurs'])])
        ws_resume.append(["⚠️ Alertes", len(s['alertes'])])
        
        # ===== ONGLET 2 : SOURCES + CAMEMBERT =====
        ws_src = wb.create_sheet("🔵 Sources")
        ws_src.column_dimensions['A'].width = 35
        ws_src.column_dimensions['B'].width = 15
        ws_src.column_dimensions['C'].width = 12
        
        ws_src.append(["IP Source", "Paquets", "Pourcentage"])
        for cell in ws_src[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        for ip, c in s['sources']:
            pct = round(c / s['total_src'] * 100, 1) if s['total_src'] > 0 else 0
            ws_src.append([ip, c, pct])
        
        # CAMEMBERT Sources (Top 5)
        pie_src = PieChart()
        pie_src.title = "Top 5 Sources (%)"
        pie_src.height = 12
        pie_src.width = 18
        
        labels = Reference(ws_src, min_col=1, min_row=2, max_row=min(6, ws_src.max_row))
        data = Reference(ws_src, min_col=2, min_row=1, max_row=min(6, ws_src.max_row))
        pie_src.add_data(data, titles_from_data=True)
        pie_src.set_categories(labels)
        
        ws_src.add_chart(pie_src, "E2")
        
        # ===== ONGLET 3 : DESTINATIONS + CAMEMBERT =====
        ws_dst = wb.create_sheet("🔴 Destinations")
        ws_dst.column_dimensions['A'].width = 35
        ws_dst.column_dimensions['B'].width = 15
        ws_dst.column_dimensions['C'].width = 12
        
        ws_dst.append(["IP Destination", "Connexions", "Pourcentage"])
        for cell in ws_dst[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        for ip, c in s['destinations']:
            pct = round(c / s['total_dst'] * 100, 1) if s['total_dst'] > 0 else 0
            ws_dst.append([ip, c, pct])
        
        # CAMEMBERT Destinations (Top 5)
        pie_dst = PieChart()
        pie_dst.title = "Top 5 Destinations (%)"
        pie_dst.height = 12
        pie_dst.width = 18
        
        labels = Reference(ws_dst, min_col=1, min_row=2, max_row=min(6, ws_dst.max_row))
        data = Reference(ws_dst, min_col=2, min_row=1, max_row=min(6, ws_dst.max_row))
        pie_dst.add_data(data, titles_from_data=True)
        pie_dst.set_categories(labels)
        
        ws_dst.add_chart(pie_dst, "E2")
        
        # ===== ONGLET 4 : FLAGS TCP + BARRES =====
        ws_flags = wb.create_sheet("🏴 Flags TCP")
        ws_flags.column_dimensions['A'].width = 20
        ws_flags.column_dimensions['B'].width = 15
        ws_flags.column_dimensions['C'].width = 12
        
        ws_flags.append(["Flag TCP", "Paquets", "Pourcentage"])
        for cell in ws_flags[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        for flag, c in s['flags']:
            pct = round(c / s['total_flags'] * 100, 1) if s['total_flags'] > 0 else 0
            ws_flags.append([flag, c, pct])
        
        # GRAPHIQUE EN BARRES Flags
        bar_flags = BarChart()
        bar_flags.title = "Répartition des Flags TCP"
        bar_flags.y_axis.title = "Nombre de paquets"
        bar_flags.x_axis.title = "Flag"
        bar_flags.height = 12
        bar_flags.width = 18
        
        labels = Reference(ws_flags, min_col=1, min_row=2, max_row=ws_flags.max_row)
        data = Reference(ws_flags, min_col=2, min_row=1, max_row=ws_flags.max_row)
        bar_flags.add_data(data, titles_from_data=True)
        bar_flags.set_categories(labels)
        
        ws_flags.add_chart(bar_flags, "E2")
        
        # ===== ONGLET 5 : ERREURS + CAMEMBERT =====
        if s['err_types']:
            ws_err_types = wb.create_sheet("❌ Types Erreurs")
            ws_err_types.column_dimensions['A'].width = 20
            ws_err_types.column_dimensions['B'].width = 15
            
            ws_err_types.append(["Type d'erreur", "Occurrences"])
            for cell in ws_err_types[1]:
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.font = header_font
                cell.alignment = center
            
            for t, c in s['err_types'].items():
                ws_err_types.append([t, c])
            
            # CAMEMBERT Erreurs
            pie_err = PieChart()
            pie_err.title = "Répartition des erreurs (%)"
            pie_err.height = 12
            pie_err.width = 18
            
            labels = Reference(ws_err_types, min_col=1, min_row=2, max_row=ws_err_types.max_row)
            data = Reference(ws_err_types, min_col=2, min_row=1, max_row=ws_err_types.max_row)
            pie_err.add_data(data, titles_from_data=True)
            pie_err.set_categories(labels)
            
            ws_err_types.add_chart(pie_err, "D2")
            
            # Onglet détails erreurs
            ws_err_detail = wb.create_sheet("📋 Erreurs Détail")
            ws_err_detail.column_dimensions['A'].width = 10
            ws_err_detail.column_dimensions['B'].width = 15
            ws_err_detail.column_dimensions['C'].width = 80
            
            ws_err_detail.append(["Ligne", "Type", "Message"])
            for cell in ws_err_detail[1]:
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.font = header_font
                cell.alignment = center
            
            for e in s['erreurs']:
                ws_err_detail.append([e['ligne'], e['type'], e['msg']])
        
        # ===== ONGLET 6 : ALERTES =====
        if s['alertes']:
            ws_alertes = wb.create_sheet("⚠️ Alertes")
            ws_alertes.column_dimensions['A'].width = 10
            ws_alertes.column_dimensions['B'].width = 100
            
            ws_alertes.append(["N°", "Alerte"])
            for cell in ws_alertes[1]:
                cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                cell.font = header_font
                cell.alignment = center
            
            for i, a in enumerate(s['alertes'], 1):
                ws_alertes.append([i, a])
        
        # Sauvegarde
        wb.save(chemin)
        
        messagebox.showinfo(
            "✅ Export Excel réussi !",
            f"Fichier créé : {chemin}\n\n"
            f"Contient :\n"
            f"• Résumé global\n"
            f"• 3 camemberts (Sources, Destinations, Erreurs)\n"
            f"• 1 graphique en barres (Flags TCP)\n"
            f"• Tableaux détaillés avec %\n"
            f"• Liste des alertes"
        )
    
    def export_md(self):
        if not self.stats:
            return
        
        chemin = filedialog.asksaveasfilename(
            defaultextension=".md",
            filetypes=[("Fichier Markdown", "*.md")],
            initialfile="rapport_sae105.md"
        )
        
        if not chemin:
            return
        
        s = self.stats
        
        md = [
            f"# 🛡️ Rapport SAE 1.05\n\n",
            f"**📅 Date** : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n",
            f"**📁 Fichier** : `{self.fichier.name}`\n\n",
            f"---\n\n",
            f"## 📊 Statistiques globales\n\n",
            f"- **Lignes analysées** : {s['lignes']}\n",
            f"- **Sources distinctes** : {len(s['sources'])}\n",
            f"- **Destinations distinctes** : {len(s['destinations'])}\n",
            f"- **Erreurs** : {len(s['erreurs'])}\n",
            f"- **Alertes** : {len(s['alertes'])}\n\n",
            f"---\n\n",
            f"## 🔵 Top 10 Sources\n\n",
            "| Rang | IP Source | Paquets | % |\n",
            "|:----:|:----------|--------:|---:|\n"
        ]
        
        for i, (ip, c) in enumerate(s['sources'], 1):
            pct = round(c / s['total_src'] * 100, 1) if s['total_src'] > 0 else 0
            md.append(f"| {i} | `{ip}` | {c} | **{pct}%** |\n")
        
        md.append("\n---\n\n## 🔴 Top 10 Destinations\n\n")
        md.append("| Rang | IP Destination | Connexions | % |\n")
        md.append("|:----:|:---------------|----------:|---:|\n")
        
        for i, (ip, c) in enumerate(s['destinations'], 1):
            pct = round(c / s['total_dst'] * 100, 1) if s['total_dst'] > 0 else 0
            md.append(f"| {i} | `{ip}` | {c} | **{pct}%** |\n")
        
        if s['alertes']:
            md.append("\n---\n\n## ⚠️ Alertes de sécurité\n\n")
            for i, a in enumerate(s['alertes'], 1):
                md.append(f"{i}. {a}\n")
        
        md.append("\n---\n\n*Rapport généré automatiquement - SAE 1.05*\n")
        
        with open(chemin, "w", encoding="utf-8") as f:
            f.write("".join(md))
        
        messagebox.showinfo("✅ Export Markdown", f"Rapport exporté :\n{chemin}")


if __name__ == "__main__":
    if not MATPLOTLIB_OK:
        print("⚠️ matplotlib non installé. Lance : pip install matplotlib")
        print("L'app fonctionne sans graphiques pour l'instant.\n")
    
    root = tk.Tk()
    app = AnalyseurApp(root)
    root.mainloop()
